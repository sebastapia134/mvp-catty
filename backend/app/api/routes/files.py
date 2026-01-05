import copy
import json
from fastapi import APIRouter, Depends, HTTPException, status
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.api.deps import get_current_user
from app.models.file import File
from app.models.template import Template
from app.schemas.file import FileCreateIn, FileListOut, FileOut
from app.core.ids import random_code, random_share_token
from uuid import UUID

import io
import re
import unicodedata
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

router = APIRouter(prefix="/files", tags=["files"])


def _unique_file_code(db: Session) -> str:
    while True:
        code = random_code("F-", 6)
        exists = db.query(File).filter(File.code == code).first()
        if not exists:
            return code


def _unique_share_token(db: Session) -> str:
    while True:
        tok = random_share_token()
        exists = db.query(File).filter(File.share_token == tok).first()
        if not exists:
            return tok


@router.get("", response_model=list[FileListOut])
def list_files(db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    files = (
        db.query(File)
        .filter(File.owner_id == current_user.id)
        .order_by(File.updated_at.desc())
        .all()
    )
    return files


@router.post("", response_model=FileOut, status_code=201)
def create_file(payload: FileCreateIn, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    tpl = db.query(Template).filter(Template.id == payload.template_id, Template.is_active == True).first()
    if not tpl:
        raise HTTPException(status_code=404, detail="Template not found")

    if not current_user.is_admin:
        allowed = (tpl.visibility in ["public", "shared"]) or (tpl.owner_id == current_user.id)
        if not allowed:
            raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Template not allowed")

    def _unwrap_template_json(tj):
        if not isinstance(tj, dict):
            return {}

        if "template" in tj and "data" in tj and isinstance(tj["data"], dict):
            return tj["data"]

        if "data" in tj and isinstance(tj["data"], dict) and (
            "columns" in tj["data"] or "nodes" in tj["data"] or "meta" in tj["data"]
        ):
            return tj["data"]

        # Caso ideal: ya viene plano (ui/meta/columns/nodes)
        return tj

    base_json = _unwrap_template_json(copy.deepcopy(tpl.template_json))

    file_json = {
        "template": {"id": str(tpl.id), "code": tpl.code, "version": tpl.version},
        "data": base_json,
    }

    raw = json.dumps(file_json, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    size_bytes = len(raw)

    f = File(
        code=_unique_file_code(db),
        name=payload.name,
        owner_id=current_user.id,
        template_id=tpl.id,
        is_public=payload.is_public,
        share_token=_unique_share_token(db),
        share_enabled=True,
        file_json=file_json,
        size_bytes=size_bytes,
    )
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


@router.delete("/{file_id}", status_code=204)
def delete_file(file_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = db.query(File).filter(File.id == file_id).first()
    if not f:
        raise HTTPException(status_code=404, detail="File not found")

    if (not current_user.is_admin) and (f.owner_id != current_user.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    db.delete(f)
    db.commit()
    return None


def _resolve_file(db: Session, current_user, file_id: str) -> File:
    # intenta UUID primero
    uid = None
    try:
        uid = UUID(file_id)
    except Exception:
        uid = None

    q = db.query(File)
    if uid:
        q = q.filter(File.id == uid)
    else:
        q = q.filter(File.code == file_id)

    if not getattr(current_user, "is_admin", False):
        q = q.filter(File.owner_id == current_user.id)

    f = q.first()

    # si es admin y no encontró, intenta sin owner filter
    if (not f) and getattr(current_user, "is_admin", False):
        q = db.query(File)
        if uid:
            q = q.filter(File.id == uid)
        else:
            q = q.filter(File.code == file_id)
        f = q.first()

    if not f:
        raise HTTPException(status_code=404, detail="File not found")

    if (not getattr(current_user, "is_admin", False)) and (f.owner_id != current_user.id):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Not allowed")

    return f


# Mantén UNA sola definición de get_file que delega en _resolve_file
@router.get("/{file_id}", response_model=FileOut)
def get_file(file_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    return _resolve_file(db, current_user, file_id)


def _safe_filename(name: str) -> str:
    name = (name or "export").strip()
    name = re.sub(r"[^\w\-. ]+", "_", name, flags=re.UNICODE)
    name = re.sub(r"\s+", " ", name).strip()
    return name[:120] or "export"


def _unique_headers(columns: list[dict]) -> tuple[list[str], list[str], list[str]]:
    """
    Retorna (keys, headers, types) respetando el orden de columns.
    """
    used = set()
    keys, headers, types = [], [], []
    for c in columns or []:
        k = str(c.get("key", "")).strip()
        if not k:
            continue
        label = (c.get("label") or k).strip()
        h = label or k
        if h in used:
            h = f"{h} ({k})"
        used.add(h)
        keys.append(k)
        headers.append(h)
        types.append(str(c.get("type", "")).lower())
    return keys, headers, types


def _apply_header_style(ws):
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")


def _set_col_width(ws, col_idx: int, width: float):
    ws.column_dimensions[get_column_letter(col_idx)].width = width


# (Reemplaza la función _build_file_xlsx en tu files.py por este bloque)
def _build_file_xlsx(f: File) -> io.BytesIO:
    file_json = f.file_json or {}
    data = (file_json.get("data") or {}) if isinstance(file_json, dict) else {}
    tpl = (file_json.get("template") or {}) if isinstance(file_json, dict) else {}

    columns = data.get("columns") or []
    nodes = data.get("nodes") or []
    intro = data.get("intro") or []
    meta = data.get("meta") or {}
    questions = data.get("questions") or {}

    wb = Workbook()

    ws = wb.active
    ws.title = "Checklist"

    keys, headers, types = _unique_headers(columns)

    # Si no vienen columnas definidas, inferimos keys de forma determinista
    # y también incluimos las keys dentro de cada nodo.custom
    if (not keys) and isinstance(nodes, list) and nodes:
        seen = []
        seen_set = set()

        # Prioriza campos comunes en orden útil
        preferred = ["id", "code", "title", "type", "parentId", "viKey", "vcKey", "weight", "required", "active", "order"]

        # Recolecta todas las keys (incluye custom) manteniendo determinismo:
        for r in nodes:
            if not isinstance(r, dict):
                continue
            # top-level keys
            for k in r.keys():
                if k == "custom":
                    continue
                kk = str(k)
                if kk not in seen_set:
                    seen.append(kk)
                    seen_set.add(kk)
            # custom keys
            custom = r.get("custom") or {}
            if isinstance(custom, dict):
                for k in custom.keys():
                    kk = str(k)
                    if kk not in seen_set:
                        seen.append(kk)
                        seen_set.add(kk)

        # Ordena: preferred primero si existen, luego el resto en el orden observado
        ordered = []
        for p in preferred:
            if p in seen_set and p not in ordered:
                ordered.append(p)
        for s in seen:
            if s not in ordered:
                ordered.append(s)

        keys = ordered
        headers = keys
        types = [""] * len(keys)

    # Escribe encabezados y aplica estilo
    ws.append(headers)
    _apply_header_style(ws)

    # Para cada nodo, construye un mapa plano que combine top-level + custom,
    # y hace lookup tolerante (case-insensitive + normalizado)
    if isinstance(nodes, list):
        # helper para normalizar nombres de clave (quita acentos y caracteres no ascii, lower)
        def norm_str(s):
            if s is None:
                return ""
            s = str(s)
            nk = unicodedata.normalize("NFD", s)
            nk = nk.encode("ascii", "ignore").decode("ascii")
            nk = re.sub(r"\s+", "", nk).lower()
            nk = re.sub(r"[^a-z0-9_]", "", nk)  # conservamos guiones bajos por si acaso
            return nk

        # construir filas
        # Creamos varios índices de búsqueda para cada nodo: exacto, lower, upper, normalizado
        for r in nodes:
            r = r or {}
            custom = r.get("custom") or {}

            # construir flat: top-level first, luego custom (custom sobrescribe)
            flat = {}
            if isinstance(r, dict):
                for kk, vv in r.items():
                    if kk == "custom":
                        continue
                    # siempre convertir la key a str para evitar keys no-string
                    flat[str(kk)] = vv
            if isinstance(custom, dict):
                for kk, vv in custom.items():
                    flat[str(kk)] = vv

            # índices auxiliares
            flat_exact = {str(k): v for k, v in flat.items()}
            flat_lower = {str(k).lower(): v for k, v in flat.items()}
            flat_upper = {str(k).upper(): v for k, v in flat.items()}
            flat_norm = {norm_str(k): v for k, v in flat.items()}

            # también guardamos list of present normalized keys (por heurísticos)
            present_norm_keys = set(flat_norm.keys())

            row = []
            for colk in keys:
                val = None

                # prepare candidates (formas a probar)
                c0 = str(colk or "")
                cand_list = [c0, c0.lower(), c0.upper(), norm_str(c0)]

                # heurísticos adicionales: para claves numéricas, probar versiones sin/ con ceros, y si la columna suena a 'agrup' probar variaciones
                nk = norm_str(c0)
                if nk and ("agrup" in nk or "agrupa" in nk or "agr" in nk):
                    # posibles campos en nodos que contienen agrupación
                    cand_list += ["agrupacion", "agrupacion_es", "agrupacion_en", "agrupación", "agrupación_es", "agrupación_en"]
                if nk and nk.isdigit():
                    cand_list += [nk]  # ya incluido; es para claridad

                # probar candidatos en orden: exacto en flat, luego lower/upper/norm
                for c in cand_list:
                    if c in flat_exact:
                        val = flat_exact[c]
                        break
                    if c in flat_lower:
                        val = flat_lower[c]
                        break
                    if c in flat_upper:
                        val = flat_upper[c]
                        break
                    if c in flat_norm:
                        val = flat_norm[c]
                        break

                # caso especial: si buscamos parent code (columna puede llamarse PARENT_CODE o parentCode)
                if val is None and "parent" in nk:
                    # intentar resolver por parentId -> buscar nodo padre y obtener su code
                    pid = r.get("parentId") or r.get("parent") or None
                    if pid:
                        # pid puede ser uuid o número; buscamos en list nodes por id
                        parent_node = None
                        for p in nodes:
                            try:
                                if str(p.get("id")) == str(pid):
                                    parent_node = p
                                    break
                            except Exception:
                                continue
                        if parent_node:
                            val = parent_node.get("code") or parent_node.get("codigo") or ""

                # vi/vc label mapping: si columna sugiere "vi_label" o "vc_label" y el valor es clave tipo "VI_3" mapear a label si scales existen
                if isinstance(val, str) and val and (("vi" in nk and "label" in nk) or nk == "vilabel"):
                    try:
                        for s in (data.get("scales") or {}).get("VI", []) or []:
                            if s.get("key") == val:
                                val = s.get("label") or val
                                break
                    except Exception:
                        pass
                if isinstance(val, str) and val and (("vc" in nk and "label" in nk) or nk == "vclabel"):
                    try:
                        for s in (data.get("scales") or {}).get("VC", []) or []:
                            if s.get("key") == val:
                                val = s.get("label") or val
                                break
                    except Exception:
                        pass

                # fallback: si ninguna coincidencia, intentar usar propiedades estándar
                if val is None:
                    # common names
                    std_map = {
                        "id": r.get("id"),
                        "code": r.get("code") or r.get("codigo"),
                        "title": r.get("title") or r.get("titulo") or r.get("name"),
                        "desc": r.get("desc") or r.get("descripcion") or r.get("observaciones"),
                        "type": r.get("type"),
                        "vikey": r.get("viKey") or r.get("vi"),
                        "vckey": r.get("vcKey") or r.get("vc"),
                        "weight": r.get("weight"),
                        "required": r.get("required"),
                        "active": r.get("active"),
                        "order": r.get("order"),
                    }
                    # probar algunas claves comunes basadas en colk normalizado
                    if nk in ("id", "codigo", "code", "cod"):
                        val = std_map.get("id") if nk == "id" else std_map.get("code")
                    elif "title" in nk or "enunci" in nk or "nombre" in nk:
                        val = std_map.get("title")
                    elif "observ" in nk:
                        val = r.get("observaciones") or r.get("obs") or ""
                    elif "desc" in nk or "descrip" in nk:
                         val = std_map.get("desc")
                    elif "vi" in nk and "label" not in nk:
                        val = std_map.get("vikey")
                    elif "vc" in nk and "label" not in nk:
                        val = std_map.get("vckey")
                    # si sigue None, ponemos cadena vacía abajo

                # default a cadena vacía si todavía None
                if val is None:
                    val = ""

                # serializar dict/list para que openpyxl no falle
                if isinstance(val, (dict, list)):
                    try:
                        val = json.dumps(val, ensure_ascii=False)
                    except Exception:
                        val = str(val)

                row.append(val)

            # asegurar longitud
            if len(row) != len(keys):
                row = row[: len(keys)] + [""] * max(0, len(keys) - len(row))

            ws.append(row)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for idx, (k, h, t) in enumerate(zip(keys, headers, types), start=1):
        width = 14
        hlow = (h or "").lower()

        if t in ("longtext",):
            width = 90
        elif t in ("text",):
            width = 40
        elif k in ("1", "2", "3", "4", "5", "id"):
            width = 6
        elif "descrip" in hlow:
            width = 90
        elif "observ" in hlow:
            width = 60
        elif "agrup" in hlow:
            width = 30

        _set_col_width(ws, idx, width)

        if t in ("longtext",) or ("descrip" in hlow) or ("observ" in hlow) or ("justif" in hlow):
            for cell in ws[get_column_letter(idx)][1:]:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_meta = wb.create_sheet("Meta")
    ws_meta.append(["Campo", "Valor"])
    _apply_header_style(ws_meta)

    meta_rows = {
        "file_id": str(getattr(f, "id", "") or ""),
        "file_code": getattr(f, "code", "") or "",
        "file_name": getattr(f, "name", "") or "",
        "created_at": str(getattr(f, "created_at", "") or ""),
        "updated_at": str(getattr(f, "updated_at", "") or ""),
        "template_id": str(tpl.get("id", "") or ""),
        "template_code": str(tpl.get("code", "") or ""),
        "template_version": str(tpl.get("version", "") or ""),
    }

    for k, v in meta_rows.items():
        ws_meta.append([k, v])

    ws_meta.append(["", ""])

    if isinstance(meta, dict):
        for k, v in meta.items():
            ws_meta.append([k, v])

    ws_meta.column_dimensions["A"].width = 28
    ws_meta.column_dimensions["B"].width = 100
    for cell in ws_meta["B"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_q = wb.create_sheet("Preguntas")
    ws_q.append(["Key", "Text"])
    _apply_header_style(ws_q)

    if isinstance(questions, dict):
        for k, v in questions.items():
            ws_q.append([k, v])

    ws_q.column_dimensions["A"].width = 30
    ws_q.column_dimensions["B"].width = 110
    for cell in ws_q["B"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_i = wb.create_sheet("Intro")
    ws_i.append(["Index", "Text"])
    _apply_header_style(ws_i)

    if isinstance(intro, list):
        for i, t in enumerate(intro, start=1):
            ws_i.append([i, t])

    ws_i.column_dimensions["A"].width = 10
    ws_i.column_dimensions["B"].width = 110
    for cell in ws_i["B"][1:]:
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws_c = wb.create_sheet("Columnas")
    ws_c.append(["key", "label", "type"])
    _apply_header_style(ws_c)

    for c in columns or []:
        ws_c.append([c.get("key", ""), c.get("label", ""), c.get("type", "")])

    ws_c.column_dimensions["A"].width = 14
    ws_c.column_dimensions["B"].width = 70
    ws_c.column_dimensions["C"].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

@router.get("/{file_id}/export.xlsx")
def export_file_xlsx(file_id: str, db: Session = Depends(get_db), current_user=Depends(get_current_user)):
    f = _resolve_file(db, current_user, file_id)

    if not f.file_json:
        raise HTTPException(status_code=400, detail="File has no JSON to export")

    buf = _build_file_xlsx(f)

    fname = _safe_filename(f"{f.code}-{f.name}".strip("-")) + ".xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{fname}"'}

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )

@router.patch("/{file_id}", response_model=FileOut)
def update_file(
    file_id: str,
    payload: dict,
    db: Session = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """
    Actualiza file.file_json (reemplaza) y recalcula size_bytes.
    Espera payload con 'file_json' (objeto) OR 'data' (obj para poner en file_json = {'data': ...}).
    """
    f = _resolve_file(db, current_user, file_id)

    # Determinar nuevo file_json
    new_file_json = None
    if isinstance(payload, dict) and "file_json" in payload:
        new_file_json = payload["file_json"]
    elif isinstance(payload, dict) and "data" in payload:
        new_file_json = {"data": payload["data"]}
    else:
        raise HTTPException(status_code=400, detail="Payload inválido. Envía 'file_json' o 'data'.")

    # Guardamos
    try:
        raw = json.dumps(new_file_json, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
        size_bytes = len(raw)
    except Exception:
        raise HTTPException(status_code=400, detail="No se pudo serializar JSON.")

    f.file_json = new_file_json
    f.size_bytes = size_bytes
    db.add(f)
    db.commit()
    db.refresh(f)
    return f